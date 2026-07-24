import argparse
import json
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
ALT_ROW_FILL = PatternFill("solid", fgColor="EAF3F8")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F1F1F")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as source:
        scan = json.load(source)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "요약"
    create_overview(overview, scan)

    all_stocks = workbook.create_sheet("전종목수급")
    create_all_stocks(all_stocks, scan)

    institution_details = workbook.create_sheet("기관세부수급")
    create_institution_details(institution_details, scan)

    sector_summary = workbook.create_sheet("섹터별수급")
    create_sector_summary(sector_summary, scan)

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    workbook.save(args.output)
    verify(args.output, len(scan.get("records", [])))


def create_overview(ws, scan):
    compact = scan.get("scope", {}).get("type") == "compact"
    intraday = scan.get("dataMode") == "intraday-estimate"
    scope_title = "시총 2,000억원 이상" if compact else "전종목"
    mode_title = "장중 추정 수급" if intraday else "마감 수급"
    ws["A1"] = f"국내주식 {scope_title} {mode_title} 요약 — {scan['date']}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    rows = [
        ("집계 종목 수", scan.get("records", []).__len__()),
        ("대상 종목 수", scan.get("total", 0)),
        ("실패 종목 수", scan.get("failed", 0)),
        ("생성 시각", scan.get("createdAt", "")),
        ("산식", "시총 대비 = 순매수 금액 ÷ 시가총액"),
        ("데이터 구분", "장중 가집계" if intraday else "장 마감"),
        ("주의", "외국인·기관 가집계 수량 × 현재가 추정. 연기금·기관 세부는 장 마감 후 제공됩니다." if intraday else "연기금등은 KIS 투자자 수급의 fund 구분입니다."),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        ws.cell(index, 1, label).fill = SUBHEADER_FILL
        ws.cell(index, 1).font = Font(bold=True)
        ws.cell(index, 2, value)

    row = 11
    flow_groups = [("foreign", "외국인"), ("institution", "기관")] if intraday else [("foreign", "외국인"), ("pension", "연기금등"), ("institution", "기관")]
    for flow_key, label in flow_groups:
        row = write_top10_pair(ws, scan.get("records", []), flow_key, label, "absoluteWon", "절대", row)
        row = write_top10_pair(ws, scan.get("records", []), flow_key, label, "marketCapPct", "시총 대비", row)

    for column, width in {"A": 26, "B": 30, "C": 20, "D": 14, "E": 20, "F": 4, "G": 4, "H": 26, "I": 30, "J": 20, "K": 14, "L": 20}.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A3"
    return
    for flow_key, label in [("foreign", "외국인"), ("pension", "연기금등"), ("institution", "기관")]:
        row = write_top10_table(ws, scan.get("records", []), flow_key, label, "absoluteWon", "절대 순매수 금액", row)
        row = write_top10_table(ws, scan.get("records", []), flow_key, label, "marketCapPct", "시가총액 대비 수급 비중", row)

    for column, width in {"A": 26, "B": 30, "C": 20, "D": 14, "E": 20}.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A3"


def write_top10_pair(ws, records, flow_key, label, metric, metric_label, row):
    buy_next_row = write_ranked_table(ws, records, flow_key, label, metric, metric_label, row, 1, "매수", True)
    sell_next_row = write_ranked_table(ws, records, flow_key, label, metric, metric_label, row, 8, "매도", False)
    return max(buy_next_row, sell_next_row)


def write_ranked_table(ws, records, flow_key, label, metric, metric_label, row, start_col, side_label, reverse):
    ws.cell(row, start_col, f"{label} 상위 10 — {metric_label} 순{side_label} 금액" if metric == "absoluteWon" else f"{label} 상위 10 — {metric_label} 순{side_label} 비중")
    ws.cell(row, start_col).font = Font(bold=True)
    row += 1
    headers = ["순위", "종목명 (종목코드)", f"순{side_label} 금액", "시총 대비", "시가총액"]
    for offset, header in enumerate(headers):
        cell = ws.cell(row, start_col + offset, header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center")

    ranked = sorted(
        (
            record for record in records
            if record["flows"][flow_key][metric] is not None
            and (record["flows"][flow_key][metric] > 0 if reverse else record["flows"][flow_key][metric] < 0)
        ),
        key=lambda item: item["flows"][flow_key][metric],
        reverse=reverse,
    )[:10]
    for rank, record in enumerate(ranked, start=1):
        flow = record["flows"][flow_key]
        data_row = row + rank
        ws.cell(data_row, start_col, rank)
        ws.cell(data_row, start_col + 1, f'{record["name"]} ({record["code"]})')
        ws.cell(data_row, start_col + 2, flow["absoluteWon"])
        ws.cell(data_row, start_col + 3, ratio_decimal(flow["marketCapPct"]))
        ws.cell(data_row, start_col + 4, record["marketCapWon"])
        ws.cell(data_row, start_col + 2).number_format = '#,##0;[Red]-#,##0'
        ws.cell(data_row, start_col + 3).number_format = '+0.000%;[Red]-0.000%'
        ws.cell(data_row, start_col + 4).number_format = '#,##0'
    return row + len(ranked) + 3


def write_top10_table(ws, records, flow_key, label, metric, metric_label, row):
    ws.cell(row, 1, f"{label} 상위 10 — {metric_label}")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    headers = ["순위", "종목명 (종목코드)", "순매수 금액", "시총 대비", "시가총액"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    row += 1
    ranked = sorted(
        (record for record in records if record["flows"][flow_key][metric] is not None),
        key=lambda item: item["flows"][flow_key][metric],
        reverse=True,
    )[:10]
    for rank, record in enumerate(ranked, start=1):
        flow = record["flows"][flow_key]
        ws.append([rank, f'{record["name"]} ({record["code"]})', flow["absoluteWon"], ratio_decimal(flow["marketCapPct"]), record["marketCapWon"]])
    for cell in ws.iter_rows(min_row=row, max_row=row + len(ranked) - 1, min_col=3, max_col=3):
        cell[0].number_format = '#,##0;[Red]-#,##0'
    for cell in ws.iter_rows(min_row=row, max_row=row + len(ranked) - 1, min_col=4, max_col=4):
        cell[0].number_format = '+0.000%;[Red]-0.000%'
    for cell in ws.iter_rows(min_row=row, max_row=row + len(ranked) - 1, min_col=5, max_col=5):
        cell[0].number_format = '#,##0'
    return row + len(ranked) + 3


def create_all_stocks(ws, scan):
    headers = [
        "종목명 (종목코드)", "섹터", "종가", "전일대비", "등락률", "시가총액",
        "외국인 순매수금액", "외국인 순매수수량", "외국인 시총대비",
        "기관 순매수금액", "기관 순매수수량", "기관 시총대비",
        "연기금등 순매수금액", "연기금등 순매수수량", "연기금등 시총대비",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center")

    records = sorted(
        scan.get("records", []),
        key=lambda record: record.get("marketCapWon") or 0,
        reverse=True,
    )
    for record in records:
        flows = record["flows"]
        ws.append([
            f'{record["name"]} ({record["code"]})', record.get("sector", "미분류"), record["closePrice"], record.get("priceChange"), ratio_decimal(record.get("priceChangePct")), record["marketCapWon"],
            flows["foreign"]["absoluteWon"], flows["foreign"]["quantity"], ratio_decimal(flows["foreign"]["marketCapPct"]),
            flows["institution"]["absoluteWon"], flows["institution"]["quantity"], ratio_decimal(flows["institution"]["marketCapPct"]),
            flows["pension"]["absoluteWon"], flows["pension"]["quantity"], ratio_decimal(flows["pension"]["marketCapPct"]),
        ])

    last_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        row[2].number_format = '#,##0'
        row[3].number_format = '#,##0;[Red]-#,##0'
        row[4].number_format = '+0.00%;[Red]-0.00%'
        row[5].number_format = '#,##0'
        for index in [6, 7, 9, 10, 12, 13]:
            row[index].number_format = '#,##0;[Red]-#,##0'
        for index in [8, 11, 14]:
            row[index].number_format = '+0.000%;[Red]-0.000%'

    # Excel repairs openpyxl table XML when a table and worksheet AutoFilter
    # describe the same range. Keep the filter and reproduce the banded-row
    # appearance without emitting a structured-table part.
    for row_number in range(2, last_row + 1):
        if row_number % 2 == 0:
            for cell in ws[row_number]:
                cell.fill = ALT_ROW_FILL
    if last_row >= 2:
        for col in ["I", "L", "O"]:
            ws.conditional_formatting.add(f"{col}2:{col}{last_row}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    widths = [26, 24, 14, 14, 12, 20, 20, 18, 14, 20, 18, 14, 20, 18, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{last_row}"


def create_institution_details(ws, scan):
    investor_columns = [
        ("securities", "증권"),
        ("investmentTrust", "투신"),
        ("privateEquity", "사모펀드"),
        ("bank", "은행"),
        ("insurance", "보험"),
        ("merchantBank", "종금"),
        ("pensionFund", "연기금등"),
        ("otherOrganization", "기타단체"),
        ("otherCorporation", "기타법인"),
    ]
    headers = ["종목명 (종목코드)", "섹터", "종가", "전일대비", "등락률", "시가총액"]
    for _, label in investor_columns:
        headers.extend([f"{label} 순매수금액", f"{label} 순매수수량", f"{label} 시총대비"])
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center")

    records = sorted(
        scan.get("records", []),
        key=lambda record: record.get("marketCapWon") or 0,
        reverse=True,
    )
    for record in records:
        flows = record["flows"]
        row = [
            f'{record["name"]} ({record["code"]})', record.get("sector", "미분류"),
            record["closePrice"], record.get("priceChange"), ratio_decimal(record.get("priceChangePct")), record["marketCapWon"],
        ]
        for key, _ in investor_columns:
            flow = flows.get(key, {})
            row.extend([
                flow.get("absoluteWon"),
                flow.get("quantity"),
                ratio_decimal(flow.get("marketCapPct")),
            ])
        ws.append(row)

    last_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        row[2].number_format = '#,##0'
        row[3].number_format = '#,##0;[Red]-#,##0'
        row[4].number_format = '+0.00%;[Red]-0.00%'
        row[5].number_format = '#,##0'
        for column_index in range(6, len(headers), 3):
            row[column_index].number_format = '#,##0;[Red]-#,##0'
            row[column_index + 1].number_format = '#,##0;[Red]-#,##0'
            row[column_index + 2].number_format = '+0.000%;[Red]-0.000%'

    for row_number in range(2, last_row + 1):
        if row_number % 2 == 0:
            for cell in ws[row_number]:
                cell.fill = ALT_ROW_FILL
    if last_row >= 2:
        for column_index in range(9, len(headers) + 1, 3):
            column = get_column_letter(column_index)
            ws.conditional_formatting.add(
                f"{column}2:{column}{last_row}",
                ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
            )
    for column_index in range(1, len(headers) + 1):
        column = get_column_letter(column_index)
        if column_index == 1:
            ws.column_dimensions[column].width = 26
        elif column_index == 2:
            ws.column_dimensions[column].width = 24
        elif column_index in [3, 4, 5]:
            ws.column_dimensions[column].width = 14
        elif column_index >= 7 and (column_index - 9) % 3 == 0:
            ws.column_dimensions[column].width = 14
        else:
            ws.column_dimensions[column].width = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"


def create_sector_summary(ws, scan):
    sector_map = {}
    for record in scan.get("records", []):
        sector = record.get("sector") or "미분류"
        summary = sector_map.setdefault(sector, {
            "count": 0, "market_cap": 0,
            "foreign": 0, "institution": 0, "pension": 0,
            "foreign_available": False, "institution_available": False, "pension_available": False,
        })
        summary["count"] += 1
        summary["market_cap"] += record.get("marketCapWon") or 0
        for key in ["foreign", "institution", "pension"]:
            value = record["flows"][key]["absoluteWon"]
            if value is not None:
                summary[key] += value
                summary[f"{key}_available"] = True

    headers = ["섹터", "종목수", "섹터 시가총액", "외국인 순매수", "외국인 시총대비", "기관 순매수", "기관 시총대비", "연기금등 순매수", "연기금등 시총대비"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center")
    for sector, summary in sorted(sector_map.items(), key=lambda item: item[1]["market_cap"], reverse=True):
        cap = summary["market_cap"]
        foreign = summary["foreign"] if summary["foreign_available"] else None
        institution = summary["institution"] if summary["institution_available"] else None
        pension = summary["pension"] if summary["pension_available"] else None
        ws.append([
            sector, summary["count"], cap,
            foreign, foreign / cap if cap and foreign is not None else None,
            institution, institution / cap if cap and institution is not None else None,
            pension, pension / cap if cap and pension is not None else None,
        ])
    last_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        row[1].number_format = '#,##0'
        for index in [2, 3, 5, 7]:
            row[index].number_format = '#,##0;[Red]-#,##0'
        for index in [4, 6, 8]:
            row[index].number_format = '+0.000%;[Red]-0.000%'
    for row_number in range(2, last_row + 1):
        if row_number % 2 == 0:
            for cell in ws[row_number]:
                cell.fill = ALT_ROW_FILL
    if last_row >= 2:
        for col in ["E", "G", "I"]:
            ws.conditional_formatting.add(f"{col}2:{col}{last_row}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    for column, width in {"A": 28, "B": 10, "C": 20, "D": 20, "E": 14, "F": 20, "G": 14, "H": 20, "I": 14}.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{last_row}"


def verify(path, expected_rows):
    workbook = load_workbook(path, read_only=True, data_only=False)
    if workbook.sheetnames != ["요약", "전종목수급", "기관세부수급", "섹터별수급"]:
        raise RuntimeError("생성된 Excel 시트 구성이 올바르지 않습니다.")
    if workbook["전종목수급"].max_row != expected_rows + 1:
        raise RuntimeError("생성된 Excel 행 수가 수급 데이터와 일치하지 않습니다.")


def ratio_decimal(value):
    return None if value is None else value / 100


if __name__ == "__main__":
    main()
